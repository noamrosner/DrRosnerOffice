"""
    A - תאריך
    B - יום
    C - שעה
    D - שם
    E - תז
    F - מספר טלפון
    G - סוג בדיקה
    H - ביטוח
"""

import openpyxl
import pywhatkit
from tkinter.filedialog import askopenfilename

global wb
global sheet


def validInput(msg, options):
    while True:
        entered_value = input(msg)
        if entered_value in options:
            break
        else:
            print('Wrong input. Please try again')
            continue
    return entered_value


def chooseWhere():
    global sheet
    place = validInput("Ramat Hahayal or Raanana? enter 1 for RH and 2 for Raanana: ", ['1', '2'])
    if place == '1':
        place = "רמת החייל"
    if place == '2':
        place = "רעננה"

def ramatHahayalMessage(date, day, time, name, exam):
    msg = (f"\n"
           f" שלום {name},\n"
           f"זו תזכורת מהמרפאה של ד״ר רוזנר לבדיקת {exam}\n"
           f"שנקבעה לך ליום {day} {date} בשעה {time}\n"
           f"באסותא רמת החייל, רחוב הברזל 12 תל אביב קומה 5 מכון גסטרו.\n"
           f"\n"
           f"*נא להגיע חצי שעה טרם שעת הבדיקה.*\n"
           f"באם אין בכוונתך להגיע נא להודיע למספר זה בהודעה חוזרת.\n"
           f"\n"
           f"שיהיה בהצלחה,\n"
           f"בשורות טובות ובריאות שלמה ")
    return msg


def rannanaMessage(date, day, time, name, exam):
    msg = (f"\n"
           f" שלום {name},\n"
           f"זו תזכורת מהמרפאה של ד״ר רוזנר לבדיקת {exam}\n"
           f"שנקבעה לך ליום {day} {date} בשעה {time}\n"
           f"באסותא רעננה, רחוב החרושת 14 רעננה קומה 2 מכון גסטרו.\n"
           f"\n"
           f"*נא להגיע 45 דקות טרם שעת הבדיקה.*\n"
           f"באם אין בכוונתך להגיע נא להודיע למספר זה בהודעה חוזרת.\n"
           f"\n"
           f"שיהיה בהצלחה,\n"
           f"בשורות טובות ובריאות שלמה ")
    return msg

def sendMessage(phone, msg):
    try:
        pywhatkit.sendwhatmsg_instantly(phone, msg)
        print("Successfully Sent!")
    except Exception:
        print("An Unexpected Error!")


def ramatHahyalSendToList():
    global sheet
    global wb
    global filename
    for i in range(2, sheet.max_row+1):
        try:
            msg = ramatHahayalMessage(sheet[f"A{i}"].value.strftime("%d/%m/%Y"), sheet[f"B{i}"].value,
                                  sheet[f"C{i}"].value.strftime("%H:%M"), sheet[f"D{i}"].value, sheet[f"G{i}"].value)
            validPhone(f'{sheet[f"F{i}"].value}')
            sendMessage(validPhone(sheet[f"F{i}"].value), msg)
            sheet[f"I{i}"].value = "Sent successfully"
        except Exception:
            sheet[f"I{i}"].value = "NOT SENT"
    wb.save(filename)

def rannanaSendToList():
    global sheet
    global wb
    global filename
    for i in range(2, sheet.max_row+1):
        try:
            msg = rannanaMessage(sheet[f"A{i}"].value.strftime("%d/%m/%Y"), sheet[f"B{i}"].value,
                                  sheet[f"C{i}"].value.strftime("%H:%M"), sheet[f"D{i}"].value, sheet[f"G{i}"].value)
            validPhone(f'{sheet[f"F{i}"].value}')
            sendMessage(validPhone(sheet[f"F{i}"].value), msg)
            sheet[f"I{i}"].value = "Sent successfully"
        except Exception:
            sheet[f"I{i}"].value = "NOT SENT"
    wb.save(filename)

def validPhone(num):
    if num[0] == "0":
        num = num[1:]
    phone = num.replace('-','')
    return f"+972{phone}"

def ramatHahyalAskFile():
    global sheet
    global wb
    global filename
    while True:
        try:
            filename = askopenfilename()
            wb = openpyxl.load_workbook(f'{filename}')
            sheet = wb['רמת החייל']
            print(f"Opening {filename}!\n")
            print(
                "\n--------------------------------------------------------------------------------------------------------"
                "-\n")
            break
        except openpyxl.utils.exceptions.InvalidFileException:
            print("Fix this")
        except NameError:
            print("Wrong file, not a xlsx file")
        except TypeError:
            print("Wrong file, not a xlsx file")


def rannanaAskFile():
    global sheet
    global wb
    global filename
    while True:
        try:
            filename = askopenfilename()
            wb = openpyxl.load_workbook(f'{filename}')
            sheet = wb['רעננה']
            print(f"Opening {filename}!\n")
            print(
                "\n--------------------------------------------------------------------------------------------------------"
                "-\n")
            break
        except openpyxl.utils.exceptions.InvalidFileException:
            print("Fix this")
        except NameError:
            print("Wrong file, not a xlsx file")
        except TypeError:
            print("Wrong file, not a xlsx file")


def ramatHahyal():
    ramatHahyalAskFile()
    ramatHahyalSendToList()

def rannana():
    rannanaAskFile()
    rannanaSendToList()

if __name__ == "__main__":
    #ramatHahyal()
    rannana()
